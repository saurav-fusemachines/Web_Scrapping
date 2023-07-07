import os 
import requests
import re
from selenium import webdriver
import openpyxl
import yaml
import warnings
from scrapy.selector import Selector
from time import sleep
from webdriver_manager.firefox import GeckoDriverManager 

from selenium.webdriver.common.keys import Keys

from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC  
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from thefuzz import fuzz
from openpyxl import Workbook
import datetime

def check_exists_by_xpath(driver, xpath_expression):
    try:
        driver.find_element(By.XPATH,xpath_expression).click()
    except NoSuchElementException:
        return False
    return True

# def get_wb_obj(sheet_name):
#     # wb_obj = openpyxl.load_workbook('/home/fm-pc-lt-316/FuseMachines/IA_AWS/ML_Job_Scrape/Scraped_Jobs.xlsx')
#     # sheet_obj = wb_obj[sheet_name]
#     # return wb_obj,sheet_obj

def create_file(config):
    filename = str(datetime.datetime.today().date())
    wb = Workbook()
    ws = wb.create_sheet(f'{list(config.keys())[0].title()}')
    sheet_names = wb.sheetnames
    print(sheet_names)
    # wb.remove(sheet_names[0])
    ws.cell(row=1,column=1).value='Search_Keyword_for_Job'
    ws.cell(row=1, column=2).value = 'URL'
    ws.cell(row=1,column=3).value='Position'
    ws.cell(row=1,column=4).value='Company'
    ws.cell(row=1,column=5).value='Job Location'
    ws.cell(row=1, column=6).value='Job Locatino Type'
    ws.cell(row=1,column=7).value='Salary Range'
    ws.cell(row=1,column=8).value='Job Type'
    ws.cell(row=1,column=9).value='Job Description'
    wb.save(f'./{filename}.xlsx')
    return wb, ws,filename



# Open the YAML file
with open(os.path.join(os.getcwd(), 'jobs.yml'), 'r') as yaml_file:
    # Load the YAML content
    config = yaml.safe_load(yaml_file)
    

driver = webdriver.Chrome()
driver.maximize_window()

row = 2
for job in config['indeed']['jobs']:
    scrape = True
    driver.get(config['indeed']['url'])
    driver.find_element(By.XPATH, "//div[@class='icl-TextInput']").click()
    what = driver.find_element(By.XPATH,"//input[@id='text-input-what']")
    what.clear()
    what.send_keys(job)
    sleep(1)
    where = driver.find_element(By.XPATH,"//input[@id='text-input-where']")
    where.send_keys(Keys.CONTROL+'a')
    where.send_keys(Keys.DELETE)
    where.send_keys(config['indeed']['location'])
    sleep(1)

    driver.find_element(By.XPATH, "//button[@class='yosegi-InlineWhatWhere-primaryButton']").click()
    
    driver.find_element(By.XPATH, "//button[@id='filter-dateposted']").click()
    sleep(1)
    driver.find_element(By.XPATH,"//li[@class='yosegi-FilterPill-dropdownListItem'][3]").click()
    while scrape:
        count= 0
        sel = Selector(text=driver.page_source) 
        cards = sel.xpath("//ul[@class='jobsearch-ResultsList css-0']/li")
        wb,ws,filename = create_file(config)
        # wb = Workbook()
        # ws = wb.create_sheet(f'{list(config.keys())[0].title()}')
    
        # ws.cell(row=1,column=1).value='Search_Keyword_for_Job'
        # ws.cell(row=1, column=2).value = 'URL'
        # ws.cell(row=1,column=3).value='Position'
        # ws.cell(row=1,column=4).value='Company'
        # ws.cell(row=1,column=5).value='Job Location'
        # ws.cell(row=1, column=6).value='Job Locatino Type'
        # ws.cell(row=1,column=7).value='Salary Range'
        # ws.cell(row=1,column=8).value='Job Type'
        # ws.cell(row=1,column=9).value='Job Description'
        
        for card in cards:
            job_position = card.xpath('.//div[@class="css-1m4cuuf e37uo190"]/h2/a/span/@title').extract_first()
            if fuzz.partial_ratio(job, job_position) < 75 or job_position is None:
                continue
            divs = card.xpath('.//div[@class="heading6 tapItem-gutter metadataContainer noJEMChips salaryOnly"]/div')
            if len(divs)==1:
                if divs.xpath('.//svg/@aria-label').extract_first()=='Salary':
                    salary = divs.xpath('.//text()').extract_first()
                    job_type=''
                if divs.xpath('.//svg/@aria-label').extract_first()=='Job type':
                    job_type = divs.xpath('.//text()').extract_first()
                    salary = ''
            elif len(divs)==2:
                if divs[0].xpath('./@class').extract_first()=='metadata estimated-salary-container':
                    salary = divs.xpath('.//span/text()').extract_first()
                else:
                    salary = divs.xpath(".//div/text()").extract_first()
                job_type = divs[1].xpath('.//div/text()').extract_first()
            else:
                job_type=''
                salary = ''
            try:
                job_link = card.xpath('.//div[@class="css-1m4cuuf e37uo190"]/h2/a/@href').extract_first()
                match = re.search(r"jk=([^&]+)", job_link)

                if match:
                    result = match.group(1)
                    
                url = f'https://indeed.com/viewjob?jk={result}'
                # print(url)
            except Exception as e:
                pass
            driver.execute_script("window.open('');")
            driver.switch_to.window(driver.window_handles[1]) 
            driver.get(url)
            sleep(5)
            job_sec_full_page_src = driver.page_source
            job_sel = Selector(text=job_sec_full_page_src)
            # print(job_sel)
            # position = job_sel.xpath("//div[@class='jobsearch-JobInfoHeader-title-container ']/h1/span/text()").extract_first()
            # print(position)
            # if fuzz.partial_ratio(job, position) < 75 or position is None:
            #     continue
            company = job_sel.xpath("//div[@data-company-name='true']/text()").extract_first()
            # print(company)
            location = job_sel.xpath("//div[@class='css-39gvaf eu4oa1w0']/div[2]/div/text()").extract_first()
            job_location = job_sel.xpath("//div[@class='css-39gvaf eu4oa1w0']/div[3]/div/text()").extract_first()
            # salary_range = job_sel.xpath("//div[@id='salaryInfoAndJobType']/span/text()").extract_first()
            # try:
            #     job_type = job_sel.xpath("//div[@id='salaryInfoAndJobType']/span[2]/text()").extract()[1]
            # except:
            #     job_type=''
            job_description = ''
            try:
                job_description_list = job_sel.xpath("//div[@id='jobDescriptionText']/text()").extract()
                if len(job_description_list)==0:
                    job_description_list = job_sel.xpath("//div[@id='jobDescriptionText']/div/p/text()").extract() 
                if len(job_description_list)==0:
                    job_description_list = job_sel.xpath("//div[@id='jobDescriptionText']/div/div/text()").extract() 
                # job_description_list = job_sel.xpath("//div[@id='jobDescriptionText']/div/div/text()").extract() 
            except:
                job_description
            cleaned_description=''.join(job_description_list).replace('\n','')
            print(f'Position: {job_position}\nCompany:{company}\nlocation:{location}\njob_location:{job_location}\nsalary_range:{salary}\n')
            print(f'URL: {url}')
            print(f'Job Type: {job_type}')
            print(f'Description: {cleaned_description}')
            #writing values
            ws.cell(row=row,column=1).value=job
            ws.cell(row=row, column=2).value = url
            ws.cell(row=row,column=3).value=job_position
            ws.cell(row=row,column=4).value=company
            ws.cell(row=row,column=5).value=location
            ws.cell(row=row, column=6).value=job_location
            ws.cell(row=row,column=7).value=salary
            ws.cell(row=row,column=8).value=job_type
            ws.cell(row=row,column=9).value=cleaned_description
            wb.save(f'./{filename}.xlsx')
            row+=1
            count+=1
            print(count)
            print('---------------------------------')
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
        #     if count == 2:
        #         break
        # scrape=False
        scrape = check_exists_by_xpath(driver, "//a[@aria-label='Next Page']")
    # break
        
